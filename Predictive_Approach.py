import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import time
from tqdm import tqdm
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import MinMaxScaler
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

start_time=time.time()

input_file="AutoFinance_Extended_Jan2022_to_Sep2025.xlsx"
output_dir="output"
os.makedirs(output_dir,exist_ok=True)

pd.options.display.float_format='{:.0f}'.format

forecast_horizon=24

xls=pd.ExcelFile(input_file)
sheets=xls.sheet_names

styles=getSampleStyleSheet()

styles.add(ParagraphStyle(name="MainTitleStyle",alignment=1,fontSize=26,leading=30))
styles.add(ParagraphStyle(name="SectionHeadingStyle",alignment=1,fontSize=18,leading=22))
styles.add(ParagraphStyle(name="BodyTextStyleCustom",fontSize=11,leading=16))

def forecast_series(series):

    scaler=MinMaxScaler()

    scaled=scaler.fit_transform(series.reshape(-1,1)).flatten()

    X=[]
    y=[]

    lag=6

    for i in range(lag,len(scaled)):
        X.append(scaled[i-lag:i])
        y.append(scaled[i])

    if len(X)==0:
        return np.repeat(series[-1],forecast_horizon)

    model=RandomForestRegressor(n_estimators=500,random_state=42)

    model.fit(X,y)

    window=scaled[-lag:]

    preds=[]

    for _ in range(forecast_horizon):

        pred=model.predict(window.reshape(1,-1))[0]

        preds.append(pred)

        window=np.append(window[1:],pred)

    preds=scaler.inverse_transform(np.array(preds).reshape(-1,1)).flatten()

    return preds

writer=pd.ExcelWriter(os.path.join(output_dir,"Forecast_Output.xlsx"),engine="openpyxl")

total_rows=sum(pd.read_excel(input_file,sheet_name=s).shape[0] for s in sheets)

progress=tqdm(total=total_rows,desc="Forecast Progress")

for sheet in sheets:

    df=pd.read_excel(input_file,sheet_name=sheet)

    metric_col=df.iloc[:,0]
    tier_col=df.iloc[:,1]
    data=df.iloc[:,2:]

    hist_dates=pd.to_datetime(data.columns,format="%b-%y")

    future_dates=[hist_dates[-1]+relativedelta(months=i+1) for i in range(forecast_horizon)]

    forecast_data=[]

    explanations=[]

    sheet_dir=os.path.join(output_dir,sheet)

    os.makedirs(sheet_dir,exist_ok=True)

    for idx,row in data.iterrows():

        series=row.values.astype(float)

        preds=forecast_series(series)

        forecast_data.append(np.round(preds))

        metric=metric_col.iloc[idx]
        tier=tier_col.iloc[idx]

        deviation=np.std(series)

        growth=((series[-1]-series[0])/series[0])*100 if series[0]!=0 else 0

        plt.figure(figsize=(12,6))
        plt.plot(hist_dates,series,label="Historical",linewidth=2)
        plt.plot(future_dates,preds,label="Forecast",linestyle="--",linewidth=2)
        plt.title(f"{sheet} | {metric} | {tier} | Historical vs Forecast")
        plt.xlabel("Month-Year")
        plt.ylabel(metric)
        plt.legend()
        plt.grid()
        image_path=os.path.join(sheet_dir,f"{metric}_{tier}_trend.png")
        plt.savefig(image_path,dpi=150)
        plt.close()

        trend_explanation=f"""
Image Type: Historical vs Forecast Trend

Metric: {metric}
Tier: {tier}

Definition:
Displays historical operational financial data and predicted continuation.

Baseline Period:
{hist_dates[0].strftime('%B %Y')} to {hist_dates[-1].strftime('%B %Y')}

Forecast Period:
{future_dates[0].strftime('%B %Y')} to {future_dates[-1].strftime('%B %Y')}

Observed Deviation:
{deviation:.2f}

Observed Growth:
{growth:.2f}%

Meaning:
Forecast extends historical trend using machine learning time-series modeling.

Hidden Insight:
Trend slope indicates operational performance strength and expected continuation.

Origin:
Derived directly from historical operational financial records.
"""

        explanations.append((image_path,trend_explanation))

        full_dates=list(hist_dates)+future_dates
        full_series=np.concatenate([series,preds])

        plt.figure(figsize=(12,6))
        plt.plot(full_dates,full_series,linewidth=2,color="blue")
        plt.axvline(hist_dates[-1],color="red",linestyle="--",linewidth=2)
        plt.title(f"{sheet} | {metric} | {tier} | Full Time Series Forecast")
        plt.xlabel("Month-Year")
        plt.ylabel(metric)
        plt.grid()
        timeseries_path=os.path.join(sheet_dir,f"{metric}_{tier}_timeseries.png")
        plt.savefig(timeseries_path,dpi=150)
        plt.close()

        timeseries_explanation=f"""
Image Type: Full Time Series Forecast

Metric: {metric}
Tier: {tier}

Definition:
Continuous time-series visualization combining historical baseline and forecast horizon.

Red Vertical Line:
Represents transition from historical observed data to predicted forecast data.

Meaning:
Shows complete temporal evolution including learned historical pattern and projected future trend.

Hidden Insight:
Stable slope indicates consistent operational performance.
Increasing slope indicates growth acceleration.
Irregular variation indicates volatility risk.

Origin:
Generated using time-series machine learning trained on historical operational data.
"""

        explanations.append((timeseries_path,timeseries_explanation))

        progress.update(1)

    forecast_df=pd.DataFrame(
        np.array(forecast_data,dtype=np.int64),
        columns=[d.strftime("%b-%y") for d in future_dates]
    )

    data_fixed=data.copy()

    for col in data_fixed.columns:
        try:
            data_fixed[col]=pd.to_numeric(data_fixed[col])
        except Exception:
            pass

    final=pd.concat([metric_col,tier_col,data_fixed,forecast_df],axis=1)

    final.to_excel(writer,sheet_name=sheet,index=False)

    pdf_path=os.path.join(output_dir,f"{sheet}_Forecast_Report.pdf")

    doc=SimpleDocTemplate(pdf_path)

    elements=[]

    elements.append(Paragraph(f"{sheet} Financial Forecast Intelligence Report",styles["MainTitleStyle"]))

    executive_summary=f"""
Executive Summary

This report analyzes Applications, Approvals, Funded Units, Originations, and Net Spread metrics.

Forecast generated using Random Forest time-series machine learning model.

Baseline Period:
{hist_dates[0].strftime('%B %Y')} to {hist_dates[-1].strftime('%B %Y')}

Forecast Horizon:
{future_dates[0].strftime('%B %Y')} to {future_dates[-1].strftime('%B %Y')}

Interpretation:
Forecast is derived purely from historical operational financial performance.

Confidence Level:
Based on historical trend stability and structural consistency.
"""

    elements.append(Spacer(1,20))
    elements.append(Paragraph(executive_summary,styles["BodyTextStyleCustom"]))
    elements.append(PageBreak())

    for img,txt in explanations:

        elements.append(Image(img,width=6.5*inch,height=4.5*inch))
        elements.append(Spacer(1,12))
        elements.append(Paragraph(txt,styles["BodyTextStyleCustom"]))
        elements.append(PageBreak())

    doc.build(elements)

writer.close()

wb=load_workbook(os.path.join(output_dir,"Forecast_Output.xlsx"))

for sheet in wb.sheetnames:

    ws = wb[sheet]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

        metric = row[0].value

        for cell in row[2:]:

            if metric == "Originations (INR)":
                cell.number_format = '#,##0'

            elif metric == "Net Spread (%)":
                cell.number_format = '0.0000'

            else:
                cell.number_format = '0'

wb.save(os.path.join(output_dir,"Forecast_Output.xlsx"))

progress.close()

end_time=time.time()

print("Complete")
print(f"Time Taken: {end_time-start_time:.2f} seconds")
print(f"Output saved in: {output_dir}")